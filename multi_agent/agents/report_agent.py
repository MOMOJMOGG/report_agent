"""
Report Agent for generating comprehensive Excel reports from insights and data.
Creates multi-tab Excel files with executive summaries, detailed analysis, and visualizations.
"""

import asyncio
import os
from datetime import datetime, date
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass
import pandas as pd
import numpy as np
from pathlib import Path

from src.main.python.core.base_agent import BaseAgent, AgentConfig
from src.main.python.models.message_types import (
    BaseMessage, MessageType, AgentType, 
    InsightsPayload, InsightData, ReportData, ReportReadyPayload, create_message
)


@dataclass
class ReportConfig:
    """Configuration for report generation."""
    # Output settings
    output_directory: str = "output/reports"
    file_prefix: str = "retail_analysis"
    include_timestamp: bool = True
    
    # Excel settings
    create_charts: bool = True
    auto_adjust_columns: bool = True
    freeze_panes: bool = True
    
    # Content settings
    include_executive_summary: bool = True
    include_detailed_analysis: bool = True
    include_raw_data: bool = True
    include_ai_insights: bool = True
    
    # Chart settings
    chart_style: str = "colorful"  # colorful, monochrome, corporate
    max_chart_categories: int = 10
    
    # Data filtering
    min_confidence_threshold: float = 0.5
    max_insights_per_category: int = 20


class ReportAgent(BaseAgent):
    """
    Agent responsible for generating comprehensive Excel reports.
    Transforms insights and data into professional business reports.
    """
    
    def __init__(self, config: Optional[AgentConfig] = None, report_config: Optional[ReportConfig] = None):
        # Configure for report generation operations
        agent_config = config or AgentConfig(
            max_retries=3,
            retry_delay=2.0,
            timeout_seconds=600,  # Longer timeout for file generation
            heartbeat_interval=30
        )
        
        super().__init__(AgentType.REPORT, agent_config, "ReportAgent")
        
        # Report configuration
        self.report_config = report_config or ReportConfig()
        
        # Register message handlers
        self.register_handler(MessageType.INSIGHTS, self.handle_insights)
        
        # Ensure output directory exists
        Path(self.report_config.output_directory).mkdir(parents=True, exist_ok=True)
        
        # Report generation statistics
        self.reports_generated = 0
        self.total_file_size = 0
    
    async def _on_start(self):
        """Initialize report agent."""
        self.logger.info("Report agent started")
        self.logger.info(f"Output directory: {self.report_config.output_directory}")
        
        # Check if openpyxl is available for Excel generation
        try:
            import openpyxl
            self.excel_available = True
            self.logger.info("Excel generation available with openpyxl")
        except ImportError:
            self.excel_available = False
            self.logger.warning("openpyxl not available - will generate CSV reports instead")
    
    async def _on_stop(self):
        """Cleanup report agent."""
        self.logger.info(f"Report agent stopped. Generated {self.reports_generated} reports, "
                        f"total size: {self.total_file_size / 1024:.1f} KB")
    
    async def handle_insights(self, message: BaseMessage) -> BaseMessage:
        """
        Handle INSIGHTS message and generate reports.
        """
        try:
            # Parse message payload
            insights_payload = InsightsPayload(
                insights=[
                    InsightData(
                        text=insight["text"],
                        confidence=insight["confidence"],
                        citations=insight["citations"],
                        category=insight["category"],
                        importance=insight.get("importance")
                    ) for insight in message.payload["insights"]
                ],
                data_summaries=message.payload["data_summaries"],
                generation_metadata=message.payload["generation_metadata"]
            )
            
            self.logger.info(f"Generating reports from {len(insights_payload.insights)} insights")
            
            # Generate reports
            reports = await self._generate_reports(insights_payload)
            
            # Prepare response
            response_payload = ReportReadyPayload(
                reports=reports,
                generation_metadata={
                    "timestamp": datetime.now().isoformat(),
                    "reports_count": len(reports),
                    "total_size_bytes": sum(report.size_bytes for report in reports),
                    "excel_available": self.excel_available
                },
                summary_stats=insights_payload.data_summaries
            )
            
            # Create response message
            response = create_message(
                MessageType.REPORT_READY,
                self.agent_type,
                AgentType.DASHBOARD,  # Send to dashboard agent
                response_payload,
                message.metadata.correlation_id
            )
            
            await self.send_message(response)
            self.logger.info(f"Generated {len(reports)} reports successfully")
            
            return response
            
        except Exception as e:
            self.logger.error(f"Error handling insights: {e}")
            # Send error response
            error_response = self.create_status_message(
                AgentType.COORDINATOR,
                message.metadata.message_id,
                "failed",
                error=str(e)
            )
            await self.send_message(error_response)
            raise
    
    async def _generate_reports(self, insights_payload: InsightsPayload) -> List[ReportData]:
        """Generate comprehensive reports from insights and data."""
        reports = []
        
        # Generate main analysis report
        if self.excel_available:
            excel_report = await self._generate_excel_report(insights_payload)
            if excel_report:
                reports.append(excel_report)
        else:
            csv_reports = await self._generate_csv_reports(insights_payload)
            reports.extend(csv_reports)
        
        # Generate summary report
        summary_report = await self._generate_summary_report(insights_payload)
        if summary_report:
            reports.append(summary_report)
        
        # Update statistics
        self.reports_generated += len(reports)
        self.total_file_size += sum(report.size_bytes for report in reports)
        
        return reports
    
    async def _generate_excel_report(self, insights_payload: InsightsPayload) -> Optional[ReportData]:
        """Generate comprehensive Excel report with multiple worksheets."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import BarChart, PieChart, Reference
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if self.report_config.include_timestamp else ""
            filename = f"{self.report_config.file_prefix}_{timestamp}.xlsx" if timestamp else f"{self.report_config.file_prefix}.xlsx"
            file_path = os.path.join(self.report_config.output_directory, filename)
            
            # Create workbook
            wb = openpyxl.Workbook()
            worksheets = []
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # 1. Executive Summary
            if self.report_config.include_executive_summary:
                ws_summary = wb.create_sheet("Executive Summary")
                await self._create_executive_summary_sheet(ws_summary, insights_payload)
                worksheets.append("Executive Summary")
            
            # 2. AI Insights
            if self.report_config.include_ai_insights and insights_payload.insights:
                ws_insights = wb.create_sheet("AI Insights")
                await self._create_insights_sheet(ws_insights, insights_payload.insights)
                worksheets.append("AI Insights")
            
            # 3. Returns Analysis
            if "returns_stats" in insights_payload.data_summaries:
                ws_returns = wb.create_sheet("Returns Analysis")
                await self._create_returns_analysis_sheet(ws_returns, insights_payload.data_summaries["returns_stats"])
                worksheets.append("Returns Analysis")
            
            # 4. Warranty Analysis
            if "warranties_stats" in insights_payload.data_summaries:
                ws_warranties = wb.create_sheet("Warranty Analysis")
                await self._create_warranty_analysis_sheet(ws_warranties, insights_payload.data_summaries["warranties_stats"])
                worksheets.append("Warranty Analysis")
            
            # 5. Category Analysis
            if "products_stats" in insights_payload.data_summaries:
                ws_categories = wb.create_sheet("Category Analysis")
                await self._create_category_analysis_sheet(ws_categories, insights_payload.data_summaries["products_stats"])
                worksheets.append("Category Analysis")
            
            # 6. Data Quality Report
            if "quality_metrics" in insights_payload.data_summaries:
                ws_quality = wb.create_sheet("Data Quality")
                await self._create_data_quality_sheet(ws_quality, insights_payload.data_summaries.get("quality_metrics", {}))
                worksheets.append("Data Quality")
            
            # Save workbook
            wb.save(file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            self.logger.info(f"Excel report generated: {file_path} ({file_size} bytes)")
            
            return ReportData(
                file_path=file_path,
                report_type="excel_analysis",
                created_at=datetime.now(),
                size_bytes=file_size,
                worksheets=worksheets
            )
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            return None
    
    async def _create_executive_summary_sheet(self, ws, insights_payload: InsightsPayload):
        """Create executive summary worksheet."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Title
        ws['A1'] = "RETAIL ANALYSIS EXECUTIVE SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Report metadata
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1
        
        ws[f'A{row}'] = "Analysis Period:"
        ws[f'B{row}'] = "Last 90 days"
        row += 2
        
        # Key metrics
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        data_summaries = insights_payload.data_summaries
        
        if "returns_stats" in data_summaries:
            returns = data_summaries["returns_stats"]
            ws[f'A{row}'] = "Total Returns:"
            ws[f'B{row}'] = f"${returns.get('total_amount', 0):,.2f}"
            row += 1
            
            ws[f'A{row}'] = "Average Return Amount:"
            ws[f'B{row}'] = f"${returns.get('avg_amount', 0):,.2f}"
            row += 1
        
        if "warranties_stats" in data_summaries:
            warranties = data_summaries["warranties_stats"]
            ws[f'A{row}'] = "Total Warranty Costs:"
            ws[f'B{row}'] = f"${warranties.get('total_cost', 0):,.2f}"
            row += 1
            
            ws[f'A{row}'] = "Avg Resolution Time:"
            ws[f'B{row}'] = f"{warranties.get('avg_resolution_time', 0):.1f} days"
            row += 2
        
        # Top insights
        ws[f'A{row}'] = "KEY AI INSIGHTS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Filter high-confidence insights
        top_insights = [
            insight for insight in insights_payload.insights 
            if insight.confidence >= self.report_config.min_confidence_threshold
        ][:5]
        
        for i, insight in enumerate(top_insights, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = insight.text
            ws[f'D{row}'] = f"Confidence: {insight.confidence:.1%}"
            row += 1
        
        # Auto-adjust column widths
        if self.report_config.auto_adjust_columns:
            for col_num in range(1, 5):  # Adjust first 4 columns
                try:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_num)
                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_num)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                except Exception:
                    # Skip column adjustment if there's an error
                    pass
    
    async def _create_insights_sheet(self, ws, insights: List[InsightData]):
        """Create AI insights worksheet."""
        from openpyxl.styles import Font, Alignment
        
        # Headers
        headers = ["Category", "Insight", "Confidence", "Citations", "Importance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, insight in enumerate(insights, 2):
            ws.cell(row=row, column=1, value=insight.category.replace("_", " ").title())
            ws.cell(row=row, column=2, value=insight.text)
            ws.cell(row=row, column=3, value=f"{insight.confidence:.1%}")
            ws.cell(row=row, column=4, value=", ".join(insight.citations))
            ws.cell(row=row, column=5, value=insight.importance or 0)
        
        # Auto-adjust columns
        if self.report_config.auto_adjust_columns:
            for col_num in range(1, 6):  # Adjust first 5 columns
                try:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_num)
                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_num)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
                except Exception:
                    # Skip column adjustment if there's an error
                    pass
    
    async def _create_returns_analysis_sheet(self, ws, returns_stats: Dict[str, Any]):
        """Create returns analysis worksheet."""
        from openpyxl.styles import Font, Alignment
        
        # Title
        ws['A1'] = "RETURNS ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Summary statistics
        ws[f'A{row}'] = "Total Amount:"
        ws[f'B{row}'] = f"${returns_stats.get('total_amount', 0):,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Average Amount:"
        ws[f'B{row}'] = f"${returns_stats.get('avg_amount', 0):,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Unique Products:"
        ws[f'B{row}'] = returns_stats.get('unique_products', 0)
        row += 2
        
        # Top reasons
        if 'top_reasons' in returns_stats:
            ws[f'A{row}'] = "TOP RETURN REASONS"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Reason"
            ws[f'B{row}'] = "Count"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for reason, count in returns_stats['top_reasons'].items():
                ws[f'A{row}'] = reason
                ws[f'B{row}'] = count
                row += 1
        
        # Status distribution
        if 'status_distribution' in returns_stats:
            row += 1
            ws[f'A{row}'] = "STATUS DISTRIBUTION"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Status"
            ws[f'B{row}'] = "Count"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for status, count in returns_stats['status_distribution'].items():
                ws[f'A{row}'] = status
                ws[f'B{row}'] = count
                row += 1
    
    async def _create_warranty_analysis_sheet(self, ws, warranty_stats: Dict[str, Any]):
        """Create warranty analysis worksheet."""
        from openpyxl.styles import Font, Alignment
        
        # Title
        ws['A1'] = "WARRANTY ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Summary statistics
        ws[f'A{row}'] = "Total Cost:"
        ws[f'B{row}'] = f"${warranty_stats.get('total_cost', 0):,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Average Cost:"
        ws[f'B{row}'] = f"${warranty_stats.get('avg_cost', 0):,.2f}"
        row += 1
        
        ws[f'A{row}'] = "Avg Resolution Time:"
        ws[f'B{row}'] = f"{warranty_stats.get('avg_resolution_time', 0):.1f} days"
        row += 2
        
        # Top issues
        if 'top_issues' in warranty_stats:
            ws[f'A{row}'] = "TOP WARRANTY ISSUES"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Issue"
            ws[f'B{row}'] = "Count"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for issue, count in warranty_stats['top_issues'].items():
                ws[f'A{row}'] = issue
                ws[f'B{row}'] = count
                row += 1
    
    async def _create_category_analysis_sheet(self, ws, products_stats: Dict[str, Any]):
        """Create category analysis worksheet."""
        from openpyxl.styles import Font, Alignment
        
        # Title
        ws['A1'] = "CATEGORY ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Summary statistics
        ws[f'A{row}'] = "Total Products:"
        ws[f'B{row}'] = products_stats.get('unique_products', 0)
        row += 1
        
        ws[f'A{row}'] = "Average Price:"
        ws[f'B{row}'] = f"${products_stats.get('avg_price', 0):,.2f}"
        row += 2
        
        # Category distribution
        if 'category_distribution' in products_stats:
            ws[f'A{row}'] = "CATEGORY DISTRIBUTION"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            
            ws[f'A{row}'] = "Category"
            ws[f'B{row}'] = "Count"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for category, count in products_stats['category_distribution'].items():
                ws[f'A{row}'] = category
                ws[f'B{row}'] = count
                row += 1
    
    async def _create_data_quality_sheet(self, ws, quality_metrics: Dict[str, Any]):
        """Create data quality worksheet."""
        from openpyxl.styles import Font, Alignment
        
        # Title
        ws['A1'] = "DATA QUALITY REPORT"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Overall metrics
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = quality_metrics.get('total_records', 0)
        row += 1
        
        ws[f'A{row}'] = "Cleaned Records:"
        ws[f'B{row}'] = quality_metrics.get('cleaned_records', 0)
        row += 1
        
        ws[f'A{row}'] = "Removed Records:"
        ws[f'B{row}'] = quality_metrics.get('removed_records', 0)
        row += 1
        
        ws[f'A{row}'] = "Completion Rate:"
        ws[f'B{row}'] = f"{quality_metrics.get('completion_rate', 0):.1%}"
        row += 1
        
        ws[f'A{row}'] = "Quality Score:"
        ws[f'B{row}'] = f"{quality_metrics.get('quality_score', 0):.1%}"
        row += 1
    
    async def _generate_csv_reports(self, insights_payload: InsightsPayload) -> List[ReportData]:
        """Generate CSV reports as fallback when Excel is not available."""
        reports = []
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if self.report_config.include_timestamp else ""
            
            # Generate insights CSV
            if insights_payload.insights:
                insights_df = pd.DataFrame([
                    {
                        "category": insight.category,
                        "text": insight.text,
                        "confidence": insight.confidence,
                        "citations": ", ".join(insight.citations),
                        "importance": insight.importance or 0
                    }
                    for insight in insights_payload.insights
                ])
                
                filename = f"{self.report_config.file_prefix}_insights_{timestamp}.csv" if timestamp else f"{self.report_config.file_prefix}_insights.csv"
                file_path = os.path.join(self.report_config.output_directory, filename)
                insights_df.to_csv(file_path, index=False)
                
                file_size = os.path.getsize(file_path)
                reports.append(ReportData(
                    file_path=file_path,
                    report_type="csv_insights",
                    created_at=datetime.now(),
                    size_bytes=file_size,
                    worksheets=["insights"]
                ))
            
            # Generate summary statistics CSV
            if insights_payload.data_summaries:
                summary_data = []
                for category, stats in insights_payload.data_summaries.items():
                    if isinstance(stats, dict):
                        for key, value in stats.items():
                            summary_data.append({
                                "category": category,
                                "metric": key,
                                "value": value
                            })
                
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    filename = f"{self.report_config.file_prefix}_summary_{timestamp}.csv" if timestamp else f"{self.report_config.file_prefix}_summary.csv"
                    file_path = os.path.join(self.report_config.output_directory, filename)
                    summary_df.to_csv(file_path, index=False)
                    
                    file_size = os.path.getsize(file_path)
                    reports.append(ReportData(
                        file_path=file_path,
                        report_type="csv_summary",
                        created_at=datetime.now(),
                        size_bytes=file_size,
                        worksheets=["summary"]
                    ))
            
            self.logger.info(f"Generated {len(reports)} CSV reports")
            return reports
            
        except Exception as e:
            self.logger.error(f"Error generating CSV reports: {e}")
            return []
    
    async def _generate_summary_report(self, insights_payload: InsightsPayload) -> Optional[ReportData]:
        """Generate a text summary report."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if self.report_config.include_timestamp else ""
            filename = f"{self.report_config.file_prefix}_summary_{timestamp}.txt" if timestamp else f"{self.report_config.file_prefix}_summary.txt"
            file_path = os.path.join(self.report_config.output_directory, filename)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("RETAIL ANALYSIS SUMMARY REPORT\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Insights: {len(insights_payload.insights)}\n\n")
                
                # Key insights
                f.write("KEY INSIGHTS:\n")
                f.write("-" * 20 + "\n")
                high_confidence_insights = [
                    insight for insight in insights_payload.insights 
                    if insight.confidence >= self.report_config.min_confidence_threshold
                ][:10]
                
                for i, insight in enumerate(high_confidence_insights, 1):
                    f.write(f"{i}. {insight.text}\n")
                    f.write(f"   Confidence: {insight.confidence:.1%}\n")
                    f.write(f"   Category: {insight.category.replace('_', ' ').title()}\n\n")
                
                # Data summary
                f.write("DATA SUMMARY:\n")
                f.write("-" * 20 + "\n")
                for category, stats in insights_payload.data_summaries.items():
                    f.write(f"{category.replace('_', ' ').title()}:\n")
                    if isinstance(stats, dict):
                        for key, value in stats.items():
                            f.write(f"  {key}: {value}\n")
                    f.write("\n")
            
            file_size = os.path.getsize(file_path)
            
            return ReportData(
                file_path=file_path,
                report_type="text_summary",
                created_at=datetime.now(),
                size_bytes=file_size,
                worksheets=["summary"]
            )
            
        except Exception as e:
            self.logger.error(f"Error generating summary report: {e}")
            return None
    
    def get_report_stats(self) -> Dict[str, Any]:
        """Get report agent statistics."""
        return {
            "agent_type": self.agent_type.value,
            "agent_name": self.name,
            "reports_generated": self.reports_generated,
            "total_file_size_kb": self.total_file_size / 1024,
            "excel_available": getattr(self, 'excel_available', False),
            "output_directory": self.report_config.output_directory
        }