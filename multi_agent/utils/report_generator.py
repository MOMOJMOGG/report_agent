"""
Excel Report Generation Utility
Creates sample Excel reports with realistic data for the dashboard demo.
"""

from pathlib import Path
from datetime import datetime, timedelta
import random
from typing import Dict, List, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import uuid


class ExcelReportGenerator:
    """Generates sample Excel reports with realistic retail data."""
    
    def __init__(self, output_dir: str = "output/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_comprehensive_report(self, job_id: str) -> Dict[str, Any]:
        """Generate a comprehensive retail analysis Excel report."""
        
        filename = f"retail_analysis_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = self.output_dir / filename
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create worksheets
        self._create_executive_summary_sheet(wb)
        self._create_returns_analysis_sheet(wb)
        self._create_warranty_claims_sheet(wb)
        self._create_product_performance_sheet(wb)
        
        # Save workbook
        wb.save(file_path)
        
        return {
            "file_path": str(file_path),
            "filename": filename,
            "report_type": "excel_comprehensive",
            "created_at": datetime.now().isoformat(),
            "size_bytes": file_path.stat().st_size,
            "worksheets": ["Executive Summary", "Returns Analysis", "Warranty Claims", "Product Performance"]
        }
    
    def generate_warranty_report(self, job_id: str) -> Dict[str, Any]:
        """Generate a warranty-focused Excel report."""
        
        filename = f"warranty_deep_dive_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = self.output_dir / filename
        
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_warranty_overview_sheet(wb)
        self._create_resolution_times_sheet(wb)
        self._create_cost_analysis_sheet(wb)
        
        wb.save(file_path)
        
        return {
            "file_path": str(file_path),
            "filename": filename,
            "report_type": "excel_warranty",
            "created_at": datetime.now().isoformat(),
            "size_bytes": file_path.stat().st_size,
            "worksheets": ["Warranty Overview", "Resolution Times", "Cost Analysis"]
        }
    
    def generate_returns_report(self, job_id: str) -> Dict[str, Any]:
        """Generate a returns-focused Excel report."""
        
        filename = f"returns_pattern_analysis_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = self.output_dir / filename
        
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_return_trends_sheet(wb)
        self._create_category_analysis_sheet(wb)
        self._create_seasonal_patterns_sheet(wb)
        self._create_store_comparison_sheet(wb)
        
        wb.save(file_path)
        
        return {
            "file_path": str(file_path),
            "filename": filename,
            "report_type": "excel_returns",
            "created_at": datetime.now().isoformat(),
            "size_bytes": file_path.stat().st_size,
            "worksheets": ["Return Trends", "Category Analysis", "Seasonal Patterns", "Store Comparison"]
        }
    
    def _create_executive_summary_sheet(self, wb: Workbook):
        """Create executive summary worksheet."""
        ws = wb.create_sheet("Executive Summary")
        
        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = "Retail Analysis Executive Summary"
        ws['A1'].font = Font(bold=True, size=18)
        ws.merge_cells('A1:D1')
        
        # Key metrics
        ws['A3'] = "Key Performance Indicators"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:D3')
        
        metrics = [
            ("Total Returns Processed", "1,247"),
            ("Warranty Claims Analyzed", "432"),
            ("Return Rate", "3.2%"),
            ("Average Resolution Time", "4.2 days"),
            ("Cost Impact", "$24,580"),
            ("Customer Satisfaction Score", "4.1/5.0")
        ]
        
        for i, (metric, value) in enumerate(metrics, start=4):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # AI Insights section
        ws['A11'] = "AI-Generated Insights"
        ws['A11'].font = header_font
        ws['A11'].fill = header_fill
        ws.merge_cells('A11:D11')
        
        insights = [
            "Electronics category shows 15% higher return rate than average",
            "Returns peak during holiday season (November-December)",
            "Defective products account for 40% of warranty claims",
            "Store location significantly impacts return patterns",
            "Customer age demographic correlates with return behavior"
        ]
        
        for i, insight in enumerate(insights, start=12):
            ws[f'A{i}'] = f"• {insight}"
        
        # Set basic column widths to avoid MergedCell issues
        for col_num in range(1, 5):  # A through D
            column_letter = chr(64 + col_num)  # A=65, B=66, etc.
            ws.column_dimensions[column_letter].width = 30
    
    def _create_returns_analysis_sheet(self, wb: Workbook):
        """Create returns analysis worksheet with sample data."""
        ws = wb.create_sheet("Returns Analysis")
        
        # Sample return data
        data = []
        categories = ["Electronics", "Clothing", "Home & Garden", "Sports", "Books"]
        reasons = ["Defective", "Wrong Size", "Not as Described", "Changed Mind", "Damaged in Shipping"]
        
        for i in range(50):
            data.append({
                "Return ID": f"RET{1000 + i}",
                "Date": (datetime.now() - timedelta(days=random.randint(1, 90))).strftime("%Y-%m-%d"),
                "Category": random.choice(categories),
                "Product": f"Product {i + 1}",
                "Reason": random.choice(reasons),
                "Amount": round(random.uniform(15.99, 299.99), 2),
                "Status": random.choice(["Processed", "Pending", "Rejected"])
            })
        
        df = pd.DataFrame(data)
        
        # Write headers
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Write data
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, value in enumerate(row, 1):
                ws.cell(row=r, column=c, value=value)
        
        # Set column widths
        column_widths = [15, 12, 15, 12, 12, 10, 15]  # Adjust based on expected content
        for i, width in enumerate(column_widths, start=1):
            if i <= ws.max_column:
                column_letter = chr(64 + i)
                ws.column_dimensions[column_letter].width = width
    
    def _create_warranty_claims_sheet(self, wb: Workbook):
        """Create warranty claims worksheet."""
        ws = wb.create_sheet("Warranty Claims")
        
        # Sample warranty data
        data = []
        products = ["Laptop", "Smartphone", "Tablet", "Headphones", "Smartwatch"]
        claim_types = ["Hardware Failure", "Software Issue", "Physical Damage", "Manufacturing Defect"]
        
        for i in range(30):
            data.append({
                "Claim ID": f"WCL{2000 + i}",
                "Product": random.choice(products),
                "Purchase Date": (datetime.now() - timedelta(days=random.randint(30, 730))).strftime("%Y-%m-%d"),
                "Claim Date": (datetime.now() - timedelta(days=random.randint(1, 30))).strftime("%Y-%m-%d"),
                "Claim Type": random.choice(claim_types),
                "Status": random.choice(["Approved", "Under Review", "Rejected", "Resolved"]),
                "Cost": round(random.uniform(50, 500), 2)
            })
        
        df = pd.DataFrame(data)
        
        # Write headers
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        
        # Write data
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, value in enumerate(row, 1):
                ws.cell(row=r, column=c, value=value)
        
        # Set column widths safely
        for i in range(1, 8):  # Up to column G
            if i <= ws.max_column:
                column_letter = chr(64 + i)
                ws.column_dimensions[column_letter].width = 15
    
    def _create_product_performance_sheet(self, wb: Workbook):
        """Create product performance worksheet."""
        ws = wb.create_sheet("Product Performance")
        
        # Sample performance data
        products = [
            "Wireless Earbuds Pro", "Gaming Laptop X1", "Smart Fitness Watch",
            "4K Monitor 27\"", "Wireless Keyboard", "USB-C Hub", "Bluetooth Speaker",
            "Tablet Pro 11\"", "Smartphone Case", "Power Bank 20K"
        ]
        
        data = []
        for product in products:
            data.append({
                "Product": product,
                "Units Sold": random.randint(100, 1000),
                "Returns": random.randint(5, 50),
                "Return Rate %": round(random.uniform(1.0, 8.0), 2),
                "Warranty Claims": random.randint(2, 20),
                "Avg Rating": round(random.uniform(3.5, 5.0), 1),
                "Revenue": random.randint(10000, 100000)
            })
        
        df = pd.DataFrame(data)
        
        # Write headers
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        # Write data
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, value in enumerate(row, 1):
                ws.cell(row=r, column=c, value=value)
        
        # Set column widths safely
        for i in range(1, 8):  # Up to column G
            if i <= ws.max_column:
                column_letter = chr(64 + i)
                ws.column_dimensions[column_letter].width = 15
    
    def _create_warranty_overview_sheet(self, wb: Workbook):
        """Create warranty overview sheet."""
        ws = wb.create_sheet("Warranty Overview")
        
        ws['A1'] = "Warranty Claims Overview"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary stats
        stats = [
            ("Total Claims", "432"),
            ("Approved Claims", "389"),
            ("Rejected Claims", "43"),
            ("Approval Rate", "90.1%"),
            ("Average Processing Time", "3.2 days"),
            ("Total Cost", "$18,450")
        ]
        
        for i, (stat, value) in enumerate(stats, start=3):
            ws[f'A{i}'] = stat
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
    
    def _create_resolution_times_sheet(self, wb: Workbook):
        """Create resolution times sheet."""
        ws = wb.create_sheet("Resolution Times")
        
        # Sample resolution data
        data = []
        for i in range(20):
            data.append({
                "Claim ID": f"WCL{3000 + i}",
                "Submitted": (datetime.now() - timedelta(days=random.randint(5, 30))).strftime("%Y-%m-%d"),
                "Resolved": (datetime.now() - timedelta(days=random.randint(1, 5))).strftime("%Y-%m-%d"),
                "Days to Resolution": random.randint(1, 14),
                "Priority": random.choice(["High", "Medium", "Low"])
            })
        
        df = pd.DataFrame(data)
        
        for i, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = Font(bold=True)
        
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, value in enumerate(row, 1):
                ws.cell(row=r, column=c, value=value)
    
    def _create_cost_analysis_sheet(self, wb: Workbook):
        """Create cost analysis sheet."""
        ws = wb.create_sheet("Cost Analysis")
        
        ws['A1'] = "Warranty Cost Breakdown"
        ws['A1'].font = Font(bold=True, size=16)
        
        cost_categories = [
            ("Hardware Replacement", "$12,450"),
            ("Labor Costs", "$3,200"),
            ("Shipping & Handling", "$1,800"),
            ("Administrative Costs", "$1,000"),
            ("Total", "$18,450")
        ]
        
        for i, (category, cost) in enumerate(cost_categories, start=3):
            ws[f'A{i}'] = category
            ws[f'B{i}'] = cost
            if category == "Total":
                ws[f'A{i}'].font = Font(bold=True)
                ws[f'B{i}'].font = Font(bold=True)
    
    def _create_return_trends_sheet(self, wb: Workbook):
        """Create return trends sheet."""
        ws = wb.create_sheet("Return Trends")
        
        # Monthly trend data
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        returns = [45, 38, 52, 41, 47, 39, 58, 44, 49, 63, 78, 82]
        
        ws['A1'] = "Month"
        ws['B1'] = "Returns"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        for i, (month, return_count) in enumerate(zip(months, returns), start=2):
            ws[f'A{i}'] = month
            ws[f'B{i}'] = return_count
    
    def _create_category_analysis_sheet(self, wb: Workbook):
        """Create category analysis sheet."""
        ws = wb.create_sheet("Category Analysis")
        
        categories = ["Electronics", "Clothing", "Home & Garden", "Sports", "Books"]
        return_rates = [5.2, 8.1, 3.4, 4.7, 2.1]
        
        ws['A1'] = "Category"
        ws['B1'] = "Return Rate %"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        for i, (cat, rate) in enumerate(zip(categories, return_rates), start=2):
            ws[f'A{i}'] = cat
            ws[f'B{i}'] = rate
    
    def _create_seasonal_patterns_sheet(self, wb: Workbook):
        """Create seasonal patterns sheet."""
        ws = wb.create_sheet("Seasonal Patterns")
        
        seasons = ["Spring", "Summer", "Fall", "Winter"]
        patterns = ["Moderate returns", "Low returns", "Increasing returns", "High returns (holidays)"]
        
        ws['A1'] = "Season"
        ws['B1'] = "Pattern"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        for i, (season, pattern) in enumerate(zip(seasons, patterns), start=2):
            ws[f'A{i}'] = season
            ws[f'B{i}'] = pattern
    
    def _create_store_comparison_sheet(self, wb: Workbook):
        """Create store comparison sheet."""
        ws = wb.create_sheet("Store Comparison")
        
        stores = ["Store A", "Store B", "Store C", "Store D", "Store E"]
        return_rates = [3.2, 4.1, 2.8, 5.5, 3.7]
        
        ws['A1'] = "Store"
        ws['B1'] = "Return Rate %"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        for i, (store, rate) in enumerate(zip(stores, return_rates), start=2):
            ws[f'A{i}'] = store
            ws[f'B{i}'] = rate


def create_sample_reports():
    """Create sample reports for immediate testing."""
    generator = ExcelReportGenerator()
    
    # Generate sample reports with different job IDs
    reports = []
    
    # Comprehensive report
    job_id1 = str(uuid.uuid4())
    report1 = generator.generate_comprehensive_report(job_id1)
    reports.append(report1)
    
    # Warranty report
    job_id2 = str(uuid.uuid4())
    report2 = generator.generate_warranty_report(job_id2)
    reports.append(report2)
    
    # Returns report
    job_id3 = str(uuid.uuid4())
    report3 = generator.generate_returns_report(job_id3)
    reports.append(report3)
    
    return reports


if __name__ == "__main__":
    # Generate sample reports for testing
    reports = create_sample_reports()
    print(f"Generated {len(reports)} sample reports:")
    for report in reports:
        print(f"  - {report['filename']} ({report['report_type']})")